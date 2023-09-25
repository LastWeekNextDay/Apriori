from Apriori import Apriori
import openpyxl

excel = openpyxl.load_workbook("Input.xlsx")

page = excel.active

data = []

# Iterate the loop to read the cell values
for row in range(0, page.max_row):
    row_data = []
    for col in page.iter_cols(1, page.max_column):
        if col[row].value is not None:
            row_data.append(col[row].value)
    data.append(row_data)

apriori = Apriori()
apriori.load(data)
apriori.process(0.6, "lift", 1.0)
print(apriori.frequent_itemsets)
print(apriori.association_rules)
